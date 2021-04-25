from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from BMPConvert import Bitmap
from tkinter import Tk, Button, IntVar, Radiobutton
from tkinter.filedialog import askopenfilename

wb = Workbook()
ws = wb.active

def pixels(x,y,r,g,b):
    global ws
    cell = get_column_letter(x) + str(y*3-2)#r
    colour = "%02x0000" %r
    ws[cell].fill = PatternFill(fgColor=colour, fill_type = "solid")
    ws[cell] = r
    cell = get_column_letter(x) + str(y*3-1)#g
    colour = "00%02x00" %g
    ws[cell].fill = PatternFill(fgColor=colour, fill_type = "solid")
    ws[cell] = g
    cell = get_column_letter(x) + str(y*3)#b
    colour = "0000%02x" %b
    ws[cell].fill = PatternFill(fgColor=colour, fill_type = "solid")
    ws[cell] = b

def pixel(x,y,r,g,b):
    global ws
    cell = get_column_letter(x) + str(y)#r
    colour = "%02x%02x%02x" % (r, g, b)
    ws[cell].fill = PatternFill(fgColor=colour, fill_type = "solid")

def main():
    global file, ppp
    a = Bitmap(file)
    
    pix = a.pixels
    height = a.height
    width = a.width

    if ppp.get() == 1:
        for i in range(width):
            for j in range(height):
                pixel(i+1, j+1, pix[j][i][0], pix[j][i][1], pix[j][i][2])
            ws.column_dimensions[get_column_letter(i+1)].width = 8/3
    elif ppp.get() == 3:
        for i in range(width):
            for j in range(height):
                pixels(i+1, j+1, pix[j][i][0], pix[j][i][1], pix[j][i][2])
            ws.column_dimensions[get_column_letter(i+1)].width = 8
    else:
        print('Error!')
        exit()

    wb.save("image.xlsx")

def choosefile():
    global file
    file = askopenfilename()
    
def verifmain():
    global file
    if file != '' and file.endswith('.bmp'):
        main()

window = Tk()

file = ''
ppp = IntVar(value=1)

choose = Button(window, text='Choose File', command=choosefile)
choose.pack()
ppp1 = Radiobutton(window, text="1 cell per pixel", variable=ppp, value=1)
ppp1.pack()
ppp3 = Radiobutton(window, text="3 cells per pixel", variable=ppp, value=3)
ppp3.pack()
run = Button(window, text='Run', command=verifmain)
run.pack()

window.mainloop()